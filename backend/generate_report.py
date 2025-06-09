import sys
import pandas as pd
import json
import re
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
from azure.storage.blob import BlobServiceClient

warnings.simplefilter("ignore")  # Ignore openpyxl default style warnings

def convert_to_hours(value):
    if pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    value = str(value).strip().lower()
    if 'hour' in value:
        num = re.findall(r"[\d\.]+", value)
        return float(num[0]) if num else 0.0
    elif 'minute' in value:
        num = re.findall(r"[\d\.]+", value)
        return float(num[0])/60 if num else 0.0
    return 0.0

def download_blob_to_file(container_name, blob_name, local_path):
    connection_string = os.environ.get("AZURE_STORAGE_CONNECTION_STRING")
    if not connection_string:
        raise Exception("Missing AZURE_STORAGE_CONNECTION_STRING environment variable")

    blob_service_client = BlobServiceClient.from_connection_string(connection_string)
    blob_client = blob_service_client.get_blob_client(container=container_name, blob=blob_name)

    with open(local_path, "wb") as file:
        download_stream = blob_client.download_blob()
        file.write(download_stream.readall())

def generate_report(admin_path, activity_path, output_excel_path, output_json_path):
    admin_df = pd.read_excel(admin_path)
    activity_df = pd.read_excel(activity_path)

    admin_df['Email'] = admin_df['Email'].str.strip().str.lower()
    activity_df['Email'] = activity_df['Email'].str.strip().str.lower()

    admin_df = admin_df[['Name', 'Email', 'Program', 'License Accepted']]
    activity_df = activity_df[['Email', 'Lessons Completed', 'Video Hours Watched', 'Labs Completed']]

    # Remove entries where Program is 'LPC'
    admin_df = admin_df[admin_df['Program'].str.strip().str.upper() != 'LPC']

    merged = pd.merge(admin_df, activity_df, on='Email', how='left')

    merged['Lessons Completed'] = merged['Lessons Completed'].fillna(0).astype(int)
    merged['Video Hours Watched'] = merged['Video Hours Watched'].apply(convert_to_hours)
    merged['Labs Completed'] = merged['Labs Completed'].fillna(0).astype(int)

    def activity_status(row):
        if row['Lessons Completed'] == 0 and row['Video Hours Watched'] == 0 and row['Labs Completed'] == 0:
            return 'No activity or progress'
        return ''

    merged['Status'] = merged.apply(activity_status, axis=1)

    def license_display(val):
        if isinstance(val, str) and val.strip().lower() == 'no':
            return 'X'
        return '✓'

    merged['License Accepted Display'] = merged['License Accepted'].apply(license_display)

    final_columns = ['Name', 'Email', 'Program', 'Lessons Completed', 'Video Hours Watched', 'Labs Completed', 'License Accepted Display', 'Status']
    display_df = merged[final_columns].rename(columns={"License Accepted Display": "License Accepted"})
    display_df.to_excel(output_excel_path, index=False)

    wb = load_workbook(output_excel_path)
    ws = wb.active

    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    orange_fill = PatternFill(start_color='FFD580', end_color='FFD580', fill_type='solid')

    for row in ws.iter_rows(min_row=2):
        status = row[7].value
        license_col = row[6].value
        if status == 'No activity or progress':
            for cell in row:
                cell.fill = red_fill
        elif license_col == 'X':
            for cell in row:
                cell.fill = orange_fill

    wb.save(output_excel_path)

    json_data = display_df.to_dict(orient='records')

    # Save to root output_json_path
    with open(output_json_path, 'w') as f:
        json.dump(json_data, f, indent=2)

    # Also write to frontend public/data/ directory if available
    root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    frontend_data_dir = os.path.join(root_dir, "public", "data")
    os.makedirs(frontend_data_dir, exist_ok=True)
    frontend_json_path = os.path.join(frontend_data_dir, os.path.basename(output_json_path))
    with open(frontend_json_path, 'w') as f:
        json.dump(json_data, f, indent=2)

    # Remove backend/public/data file if it exists
    backend_data_path = os.path.join(os.path.dirname(__file__), 'public', 'data', os.path.basename(output_json_path))
    if os.path.exists(backend_data_path):
        os.remove(backend_data_path)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Uso: python generate_report.py <admin_excel_path> <activity_excel_path>")
        sys.exit(1)

    admin_xlsx = sys.argv[1]
    activity_xlsx = sys.argv[2]

    if admin_xlsx.startswith("blob:"):
        blob_name = admin_xlsx[5:]
        download_blob_to_file("kodekloud-inputs", blob_name, "admin.xlsx")
        admin_xlsx = "admin.xlsx"

    if activity_xlsx.startswith("blob:"):
        blob_name = activity_xlsx[5:]
        download_blob_to_file("kodekloud-inputs", blob_name, "activity.xlsx")
        activity_xlsx = "activity.xlsx"

    generate_report(
        admin_xlsx,
        activity_xlsx,
        "kodekloud_report.xlsx",
        "kodekloud_data.json"
    )
